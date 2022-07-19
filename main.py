# Taaha Multani @ https://github.com/taaaahahaha

import openpyxl
from bs4 import BeautifulSoup
import requests
import json
import time

def excel_scrape(path):
    print("Scraping data.......")
    li = []

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    for i in range(1,sheet_obj.max_column+1):
        # id,Asin,country
        cell_obj = sheet_obj.cell(row = 1, column = i)
        value = cell_obj.value
        if value == "id": col_id = i
        elif value == "Asin":col_asin = i
        elif value == "country":col_country = i

    start_time = time.time()

    for i in range(2,sheet_obj.max_row+1):
    
        value_id = sheet_obj.cell(row = i, column = col_id).value
        value_asin = sheet_obj.cell(row = i, column = col_asin).value
        value_country = sheet_obj.cell(row = i, column = col_country).value

        try:
            value_asin = int(value_asin)
        except:
            pass

        web_scrape(i-1,f"https://www.amazon.{value_country}/dp/{str(value_asin)}",li,int(value_id),value_country)

        if (i-1)%100 == 0:
            print("Time :" ,time.time() - start_time)
            start_time = time.time()
            print("----------------------------------------")

    
    with open("output.json", "w") as outfile:
	    json.dump(li, outfile, indent=4)




def web_scrape(sno,URL,li,id,country):

    HEADERS = ({'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
								'Accept-Language': 'en-US, en;q=0.5'})
    # Temporary Headers to avoid Amazon Bot Detection

    webpage = requests.get(URL, headers=HEADERS)


    title_tag = 'productTitle'
    price_tag = 'price'
    image_tag = 'imgTagWrapper'


    # if country=='de':
    #     pass
    # elif country=='it':
    #     title_tag = 'productTitle'
    #     price_tag = 'price'
    #     image_tag = 'imgTagWrapper'
    # elif country=='fr':
    #     pass
    # elif country=='es':
    #     pass
    # else:
    #     print(country)
    
    
    if webpage.status_code == 200:

        d = {'url' : URL}

        soup = BeautifulSoup(webpage.content, "lxml")
        try:
            title = soup.find("span", attrs={"id": title_tag})
            # print(title)
            title_value = title.string  
            title_string = title_value.strip().replace(',', '')
        except AttributeError:
            title_string = "NA"
        # print("Title = ", title_string)
        d['title'] = title_string

        
        try:
            price = soup.find("span", attrs={'id': price_tag}).string.strip().replace(',', '')
        except AttributeError:
            price = "NA"
        # print("Products price = ", price)
        d['price'] = price


        try:
            # image_url = soup.find("img", attrs={'class': 'a-dynamic-image image-stretch-vertical frontImage'})['src']
            image_url = soup.find("div", attrs={'class': image_tag}).find("img")['src']
        except AttributeError:
            image_url = "NA"
        # print("Product Image URL = ", image_url)
        d['image_url'] = image_url

        
        try:
            details = soup.find("div", attrs={'class':'a-section feature detail-bullets-wrapper bucket'}).find_all("span", attrs={'class':'a-list-item'})
        except AttributeError:
            details = "NA"
        
        # print("Details about the product : ")
        s = ""
        for i in details:
            try:
                s1,s2 = i.get_text().replace("\n","").strip().split(':')

                s1 = s1.strip()[:-37]
                s2 = s2.strip()[34:]  
                # Strip alone wasn't sufficient to remove spaces, hence used list indexing

                # print(f"{s1} = {s2}")
                s += f"{s1} = {s2} \n"
            except:
                pass
        d['details'] = s
        print(f'ID - {id}')
        print(f'SNO - {sno}',d)
        li.append(d)
        
    else:
        print(f"URL {URL} not available", webpage.status_code)

    print("---------------------------------------------------")


    



if __name__ == '__main__':
    excel_scrape('Amazon Scraping.xlsx')   # https://docs.google.com/spreadsheets/d/1BZSPhk1LDrx8ytywMHWVpCqbm8URTxTJrIRkD7PnGTM/edit#gid=0
    
    